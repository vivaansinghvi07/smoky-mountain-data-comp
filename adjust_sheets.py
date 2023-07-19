# Formats the google sheets to be more visually appealing.
# Adjusts things like column width, etc.

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

for step in ['step1', 'step2', 'step3']:
    workbook = load_workbook(f"{step}/data.xlsx")
    sheet = workbook.active
    for i, col in enumerate(sheet.iter_cols(), start=1):
        max_len = max([*map(lambda x: str(x.value if x.value is not None else ''), col)][1:], key=lambda x: len(x))
        sheet.column_dimensions[get_column_letter(i)].width = len(max_len) + 7
    workbook.save(f"{step}/data.xlsx")